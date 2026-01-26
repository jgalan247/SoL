#!/usr/bin/env python3
"""
Create a comprehensive Scheme of Learning for Computer Science Y7-Y11
Aligned with Edexcel GCSE CS 2020 Specification
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
year_fills = {
    "Y7": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Y8": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Y9": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "Y10": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Y11": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
}
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_header_style(ws, row=1):
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def apply_cell_style(cell, year=None):
    cell.alignment = Alignment(vertical='top', wrap_text=True)
    cell.border = thin_border
    if year and year in year_fills:
        cell.fill = year_fills[year]

# ============================================================================
# SHEET 1: LONG TERM PLAN (By Half-Terms)
# ============================================================================
ws_long = wb.active
ws_long.title = "Long Term Plan"

# Headers
long_headers = ["Year", "Key Stage", "Half-Term", "Topic", "Key Concepts", "GCSE Link", "Curriculum Thread", "Assessment"]
for col, header in enumerate(long_headers, 1):
    ws_long.cell(row=1, column=col, value=header)
apply_header_style(ws_long)

# Long term data - comprehensive curriculum
long_term_data = [
    # YEAR 7 - "Digital Foundations & First Steps in Programming"
    ("Y7", "KS3", "Autumn 1", "E-Safety & The Internet",
     "Online safety, social media awareness, verifying sources, digital footprint, cyberbullying, what is the internet, how the web works",
     "Topic 4: Networks, Topic 5: Issues & Impact",
     "Foundation: Understanding the digital world students will work in",
     "Online safety quiz, source verification task"),

    ("Y7", "KS3", "Autumn 2", "Binary & Data Representation",
     "Binary number system, binary to denary conversion, denary to binary, binary addition, introduction to hexadecimal, bit shifts",
     "Topic 2: Data (Binary, Hex, Data representation)",
     "Building Block: How computers store and process ALL data",
     "Binary conversion test, practical worksheet"),

    ("Y7", "KS3", "Spring 1", "Edublocks: Turtle Graphics",
     "Introduction to programming concepts through visual blocks, variables, sequence, selection (if statements), iteration (loops), turtle commands",
     "Topic 6: Problem Solving with Programming",
     "First Code: Visual introduction to programming constructs",
     "Turtle art project - create a pattern/shape"),

    ("Y7", "KS3", "Spring 2", "Programming Constructs",
     "Variables and data types, input/output, arithmetic operators, comparison operators, Boolean logic, debugging basics",
     "Topic 6: Problem Solving with Programming",
     "Core Skills: The building blocks of ALL programs",
     "Programming constructs assessment"),

    ("Y7", "KS3", "Summer 1", "Flowcharts & Python Introduction",
     "Flowchart symbols and design, algorithm representation, transitioning from blocks to Python, print, input, variables in Python",
     "Topic 1: Computational Thinking, Topic 6: Programming",
     "Design First: Planning before coding",
     "Flowchart design task, Python basics test"),

    ("Y7", "KS3", "Summer 2", "Python Quiz Project & Kodu",
     "Creating a complete Python quiz program, string manipulation, Kodu game design, event-driven programming concepts",
     "Topic 6: Problem Solving with Programming",
     "Apply & Create: First complete programming projects",
     "Python quiz project, Kodu game creation"),

    # YEAR 8 - "How Computers Think & Physical Computing"
    ("Y8", "KS3", "Autumn 1", "Logic Gates",
     "AND, OR, NOT gates, truth tables, combining gates, logic circuits, Boolean expressions, real-world applications",
     "Topic 3: Computers (Logic)",
     "Computer Thinking: How computers make decisions at hardware level",
     "Logic gate circuit design, truth table test"),

    ("Y8", "KS3", "Autumn 2", "BBC Micro:bit Programming",
     "Physical computing introduction, LEDs and buttons, sensors (accelerometer, light), variables in MakeCode/Python, loops and selection",
     "Topic 3: Embedded Systems, Topic 6: Programming",
     "Physical Code: Programming the real world",
     "Micro:bit project - interactive device"),

    ("Y8", "KS3", "Spring 1", "Algorithms: Searching & Sorting",
     "Linear search, binary search, bubble sort, merge sort (concept), algorithm efficiency, comparing algorithms",
     "Topic 1: Computational Thinking (Algorithms)",
     "Problem Solving: Standard solutions to common problems",
     "Algorithm trace tables, efficiency comparison task"),

    ("Y8", "KS3", "Spring 2", "Memory & Storage: Primary",
     "RAM vs ROM, volatile vs non-volatile, cache memory, virtual memory, how programs use memory, units of storage",
     "Topic 3: Computers (Memory)",
     "Inside the Machine: Where data lives when computing",
     "Memory concepts test, diagram labelling"),

    ("Y8", "KS3", "Summer 1", "Secondary Storage",
     "Magnetic storage (HDD), optical storage (CD/DVD), solid state (SSD/Flash), cloud storage, capacity, speed, durability, cost comparison",
     "Topic 3: Computers (Storage)",
     "Data Persistence: Where data lives permanently",
     "Storage comparison project, technology timeline"),

    ("Y8", "KS3", "Summer 2", "LEGO Robotics",
     "Robotics and automation, sensors and motors, programming robots, problem decomposition, iterative development, real-world applications",
     "Topic 1: Computational Thinking, Topic 3: Embedded Systems",
     "Automation: Programming machines to do tasks",
     "Robot challenge - complete a maze/task"),

    # YEAR 9 - "Programming Mastery & GCSE Foundations"
    ("Y9", "KS3/4", "Autumn 1", "Advanced Micro:bit Projects",
     "Multi-device communication (radio), complex sensor projects, data logging, Python on Micro:bit, project planning",
     "Topic 3: Embedded Systems, Topic 6: Programming",
     "Project Skills: Planning and building complex systems",
     "Extended Micro:bit project with documentation"),

    ("Y9", "KS3/4", "Autumn 2", "Game Development with Python",
     "Pygame basics, game loops, collision detection, user input handling, sprites and graphics, sound, game design principles",
     "Topic 6: Problem Solving with Programming",
     "Creative Coding: Applying all programming skills",
     "Complete game project"),

    ("Y9", "KS4", "Spring 1", "CPU Architecture & Fetch-Decode-Execute",
     "CPU components (ALU, CU, registers), fetch-decode-execute cycle, clock speed, cores, cache, Von Neumann architecture",
     "Topic 3: Computers (Architecture)",
     "GCSE Foundation: Understanding computer operation",
     "CPU architecture test, trace table exercise"),

    ("Y9", "KS4", "Spring 2", "Networks: Types & Topologies",
     "LAN, WAN, PAN, network topologies (star, mesh, bus), wired vs wireless, network hardware (switches, routers, NICs)",
     "Topic 4: Networks",
     "GCSE Foundation: How computers communicate",
     "Network design task, topology comparison"),

    ("Y9", "KS4", "Summer 1", "Networks: Protocols & The Internet",
     "TCP/IP, HTTP/HTTPS, FTP, email protocols, IP addresses, DNS, packet switching, internet vs web",
     "Topic 4: Networks (Protocols, Internet)",
     "GCSE Foundation: Rules of digital communication",
     "Protocol identification test, packet simulation"),

    ("Y9", "KS4", "Summer 2", "Python: Data Structures & File Handling",
     "Lists/arrays, 2D arrays, file reading and writing, CSV files, string methods, introduction to dictionaries",
     "Topic 6: Problem Solving with Programming",
     "GCSE Foundation: Managing complex data",
     "File handling project, data structure assessment"),

    # YEAR 10 - "GCSE Theory & Programming Development"
    ("Y10", "KS4", "Autumn 1", "Data Representation: Text & Images",
     "ASCII and Unicode, character sets, bitmap images, resolution, colour depth, metadata, calculating file sizes",
     "Topic 2: Data",
     "GCSE Core: How computers represent multimedia",
     "Data representation calculations test"),

    ("Y10", "KS4", "Autumn 2", "Data Representation: Sound & Compression",
     "Sound sampling, sample rate, bit depth, file sizes, lossy vs lossless compression, RLE, Huffman coding concepts",
     "Topic 2: Data (Sound, Compression)",
     "GCSE Core: Digital audio and reducing file sizes",
     "Compression comparison task, calculations test"),

    ("Y10", "KS4", "Spring 1", "Programming: Subprograms & Testing",
     "Functions and procedures, parameters, return values, local vs global scope, validation, authentication, testing strategies",
     "Topic 6: Problem Solving with Programming",
     "GCSE Core: Writing professional, modular code",
     "Modular programming project"),

    ("Y10", "KS4", "Spring 2", "Network Security & Threats",
     "Malware types, social engineering, phishing, DDoS, brute force, SQL injection, firewalls, encryption, penetration testing",
     "Topic 4: Networks (Security)",
     "GCSE Core: Protecting systems and data",
     "Security threats poster, encryption task"),

    ("Y10", "KS4", "Summer 1", "Ethical, Legal & Environmental Issues",
     "Data Protection Act, Computer Misuse Act, Copyright, Creative Commons, environmental impact, e-waste, digital divide, privacy",
     "Topic 5: Issues and Impact",
     "GCSE Core: Responsible computing",
     "Case study analysis, legislation quiz"),

    ("Y10", "KS4", "Summer 2", "NEA Preparation & Programming Practice",
     "NEA requirements, programming task practice, debugging complex programs, documentation skills, exam-style questions",
     "Topic 6: Problem Solving + NEA",
     "Exam Preparation: Getting ready for assessments",
     "Mock NEA task, programming assessment"),

    # YEAR 11 - "GCSE Completion & Exam Preparation"
    ("Y11", "KS4", "Autumn 1", "NEA Completion",
     "Complete NEA programming task, testing and refinement, documentation, code annotation, evaluation",
     "NEA (20% of GCSE)",
     "Coursework: Major programming assessment",
     "NEA submission"),

    ("Y11", "KS4", "Autumn 2", "Computational Thinking & Algorithms Review",
     "Decomposition, abstraction, pattern recognition, algorithm design, pseudocode, trace tables, efficiency analysis",
     "Topic 1: Computational Thinking",
     "Revision: Problem-solving approaches",
     "Past paper questions, mock assessment"),

    ("Y11", "KS4", "Spring 1", "Systems Architecture & Memory Review",
     "CPU architecture deep dive, memory hierarchy, storage technologies, embedded systems, exam technique",
     "Topic 3: Computers",
     "Revision: Hardware and systems",
     "Topic test, past paper questions"),

    ("Y11", "KS4", "Spring 2", "Networks & Security Review",
     "Complete network revision, security measures, protocols, exam-style scenario questions",
     "Topic 4: Networks",
     "Revision: Networks and security",
     "Full topic assessment"),

    ("Y11", "KS4", "Summer 1", "Final Revision & Exam Preparation",
     "All topics revision, past papers, exam technique, timed practice, weak area focus",
     "All Topics",
     "Exam Ready: Final preparation",
     "Mock exams, past papers"),

    ("Y11", "KS4", "Summer 2", "GCSE Examinations",
     "Paper 1: Principles of Computer Science, Paper 2: Application of Computational Thinking",
     "Final Assessment",
     "Examinations",
     "GCSE Exams"),
]

# Populate long term data
for row_idx, data in enumerate(long_term_data, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_long.cell(row=row_idx, column=col_idx, value=value)
        apply_cell_style(cell, data[0])

# Set column widths
long_widths = [6, 10, 12, 35, 70, 40, 45, 35]
for col, width in enumerate(long_widths, 1):
    ws_long.column_dimensions[get_column_letter(col)].width = width

# Freeze header row
ws_long.freeze_panes = 'A2'

# ============================================================================
# SHEET 2: MEDIUM TERM PLAN (By Weeks)
# ============================================================================
ws_medium = wb.create_sheet("Medium Term Plan")

# Headers
medium_headers = ["Year", "Half-Term", "Week", "Topic", "Learning Objectives", "Key Terms", "Resources", "Homework"]
for col, header in enumerate(medium_headers, 1):
    ws_medium.cell(row=1, column=col, value=header)
apply_header_style(ws_medium)

# Medium term data (week by week breakdown)
medium_term_data = [
    # YEAR 7 - AUTUMN 1: E-Safety & The Internet
    ("Y7", "Autumn 1", 1, "What is the Internet?", "Understand the difference between internet and web; Identify how devices connect to the internet", "Internet, World Wide Web, ISP, Router, Server", "Internet simulation activity, network diagram", "Research: History of the internet"),
    ("Y7", "Autumn 1", 2, "Online Safety Fundamentals", "Identify online risks; Understand personal information protection", "Cyberbullying, Personal data, Privacy settings, Stranger danger", "CEOP resources, case studies", "Create online safety poster"),
    ("Y7", "Autumn 1", 3, "Social Media Awareness", "Evaluate social media benefits and risks; Understand digital footprint", "Digital footprint, Reputation, Terms of service, Age restrictions", "Social media case studies, footprint checker", "Audit own digital footprint"),
    ("Y7", "Autumn 1", 4, "Verifying Online Sources", "Evaluate website reliability; Identify fake news and misinformation", "Bias, Primary source, Fake news, Fact-checking, Domain types", "Source comparison activity, fake news quiz", "Find 3 reliable sources on a topic"),
    ("Y7", "Autumn 1", 5, "Cyberbullying & Digital Citizenship", "Define cyberbullying; Know how to respond; Be a positive digital citizen", "Bystander, Upstander, Report, Block, Screenshot evidence", "Role-play scenarios, reporting practice", "Design anti-cyberbullying campaign"),
    ("Y7", "Autumn 1", 6, "E-Safety Assessment & Review", "Demonstrate e-safety knowledge; Apply learning to scenarios", "All previous terms", "Assessment materials", "Prepare for next topic"),

    # YEAR 7 - AUTUMN 2: Binary & Data Representation
    ("Y7", "Autumn 2", 1, "Introduction to Binary", "Understand why computers use binary; Convert small binary to denary", "Binary, Denary, Bit, Base 2, Base 10", "Binary cards, place value chart", "Binary conversion practice"),
    ("Y7", "Autumn 2", 2, "Binary to Denary Conversion", "Convert 8-bit binary numbers to denary accurately", "Byte, Place value, Powers of 2", "Conversion worksheets, online converter", "10 binary to denary conversions"),
    ("Y7", "Autumn 2", 3, "Denary to Binary Conversion", "Convert denary numbers to 8-bit binary", "Remainder method, Subtraction method", "Practice worksheets", "10 denary to binary conversions"),
    ("Y7", "Autumn 2", 4, "Binary Addition", "Add two 8-bit binary numbers; Understand overflow", "Carry, Overflow, Sum", "Addition worksheets, visual aids", "Binary addition problems"),
    ("Y7", "Autumn 2", 5, "Introduction to Hexadecimal", "Understand hexadecimal; Convert between hex and binary/denary", "Hexadecimal, Base 16, Nibble, A-F values", "Hex colour picker, conversion chart", "Hexadecimal conversions"),
    ("Y7", "Autumn 2", 6, "Binary Shifts & Assessment", "Perform logical shifts; Assessment on binary topics", "Left shift, Right shift, Multiply, Divide", "Shift practice, assessment paper", "Revise for assessment"),

    # YEAR 7 - SPRING 1: Edublocks: Turtle Graphics
    ("Y7", "Spring 1", 1, "Introduction to Edublocks", "Navigate Edublocks interface; Execute basic turtle commands", "Block-based coding, Sequence, Execute, Debug", "Edublocks.org, starter challenges", "Explore Edublocks at home"),
    ("Y7", "Spring 1", 2, "Sequences & Basic Shapes", "Create sequences of commands; Draw squares and triangles", "Sequence, Forward, Turn, Angle", "Shape challenge cards", "Create a house shape"),
    ("Y7", "Spring 1", 3, "Variables in Programming", "Use variables to store values; Modify variable values", "Variable, Assign, Value, Data type", "Variable experiments", "Variables practice sheet"),
    ("Y7", "Spring 1", 4, "Iteration: Loops", "Use count-controlled loops; Create repeating patterns", "Iteration, Loop, Repeat, Count-controlled", "Pattern challenges", "Design a repeating pattern"),
    ("Y7", "Spring 1", 5, "Selection: If Statements", "Use selection to make decisions; Combine with loops", "Selection, If, Condition, Boolean", "Decision-making challenges", "Selection practice"),
    ("Y7", "Spring 1", 6, "Turtle Art Project", "Apply all concepts to create artwork; Debug and refine", "Decomposition, Testing, Refinement", "Project brief, success criteria", "Complete and document project"),

    # YEAR 7 - SPRING 2: Programming Constructs
    ("Y7", "Spring 2", 1, "Variables & Data Types", "Identify data types; Declare and use variables correctly", "Integer, String, Float, Boolean, Declaration", "Data type sorting activity", "Data types worksheet"),
    ("Y7", "Spring 2", 2, "Input & Output", "Get user input; Display output; Combine input/output", "Input, Output, Print, Prompt, Concatenation", "Input/output programs", "Create input/output program"),
    ("Y7", "Spring 2", 3, "Arithmetic Operators", "Use +, -, *, /, //, %, **; Calculate expressions", "Operator, Operand, Expression, Integer division, Modulus", "Calculator challenges", "Arithmetic problems"),
    ("Y7", "Spring 2", 4, "Comparison Operators", "Use ==, !=, <, >, <=, >=; Evaluate Boolean expressions", "Comparison, Boolean, True, False, Equal", "Comparison activities", "Boolean expressions sheet"),
    ("Y7", "Spring 2", 5, "Boolean Logic", "Combine conditions with AND, OR, NOT", "AND, OR, NOT, Compound condition", "Logic puzzles", "Boolean logic practice"),
    ("Y7", "Spring 2", 6, "Debugging & Assessment", "Identify and fix errors; Assessment on constructs", "Syntax error, Logic error, Debug, Trace", "Debugging challenges, assessment", "Revise constructs"),

    # YEAR 7 - SUMMER 1: Flowcharts & Python Introduction
    ("Y7", "Summer 1", 1, "Introduction to Flowcharts", "Identify flowchart symbols; Read simple flowcharts", "Flowchart, Terminator, Process, Decision, Flow lines", "Flowchart symbol cards", "Draw daily routine flowchart"),
    ("Y7", "Summer 1", 2, "Designing Flowcharts", "Design flowcharts for simple problems; Include selection", "Algorithm, Decision, Branch, Path", "Flowchart design challenges", "Design a game flowchart"),
    ("Y7", "Summer 1", 3, "Flowcharts with Iteration", "Add loops to flowcharts; Trace flowchart execution", "Loop, Counter, Trace, Iteration", "Trace table practice", "Flowchart with loop"),
    ("Y7", "Summer 1", 4, "Introduction to Python", "Use Python IDE; Write first Python programs; print()", "IDE, Syntax, String, Print statement", "Python installation guide, first programs", "Python print challenges"),
    ("Y7", "Summer 1", 5, "Python Variables & Input", "Declare variables; Use input(); Type conversion", "Variable, input(), int(), str(), Casting", "Variable programs", "Input/output programs"),
    ("Y7", "Summer 1", 6, "Python Selection", "Write if, elif, else statements; Combine conditions", "if, elif, else, Indentation, Condition", "Selection challenges", "Selection practice programs"),

    # YEAR 7 - SUMMER 2: Python Quiz Project & Kodu
    ("Y7", "Summer 2", 1, "Python Iteration", "Write for and while loops; Use range()", "for, while, range(), Counter, Infinite loop", "Loop challenges", "Loop practice programs"),
    ("Y7", "Summer 2", 2, "Quiz Project: Planning", "Plan quiz structure; Design questions and answers", "Requirements, Design, Pseudocode, Structure", "Project planning sheet", "Research quiz questions"),
    ("Y7", "Summer 2", 3, "Quiz Project: Development", "Code the quiz; Implement scoring; Add validation", "Score, Counter, Validation, String comparison", "Development support", "Continue quiz development"),
    ("Y7", "Summer 2", 4, "Quiz Project: Completion", "Test and debug quiz; Add improvements", "Testing, Debugging, Refinement, User feedback", "Testing checklist", "Complete quiz project"),
    ("Y7", "Summer 2", 5, "Introduction to Kodu", "Navigate Kodu; Create basic world; Add characters", "Kodu, World, Character, Terrain, When-Do", "Kodu tutorials", "Explore Kodu"),
    ("Y7", "Summer 2", 6, "Kodu Game Design", "Create a complete Kodu game; Event-driven programming", "Event, Trigger, Action, Game mechanics", "Game design brief", "Complete Kodu game"),

    # YEAR 8 - AUTUMN 1: Logic Gates
    ("Y8", "Autumn 1", 1, "Introduction to Logic Gates", "Understand Boolean logic in hardware; Identify AND gate", "Logic gate, Boolean, AND gate, Truth table", "Logic gate simulator", "AND gate practice"),
    ("Y8", "Autumn 1", 2, "OR and NOT Gates", "Understand OR and NOT gates; Complete truth tables", "OR gate, NOT gate, Inverter, Input, Output", "Truth table worksheets", "OR and NOT practice"),
    ("Y8", "Autumn 1", 3, "Combining Logic Gates", "Create circuits with multiple gates; NAND, NOR, XOR awareness", "Circuit, Combination, NAND, NOR, XOR", "Circuit building activity", "Design a circuit"),
    ("Y8", "Autumn 1", 4, "Truth Tables for Circuits", "Complete truth tables for combined circuits", "Truth table, Intermediate values, Columns", "Complex truth tables", "Truth table practice"),
    ("Y8", "Autumn 1", 5, "Boolean Expressions", "Write Boolean expressions from circuits; Simplify expressions", "Expression, AND (∧), OR (∨), NOT (¬)", "Expression practice", "Write expressions"),
    ("Y8", "Autumn 1", 6, "Logic Gates Assessment", "Apply logic gate knowledge; Complete assessment", "All logic terms", "Assessment paper", "Revise logic gates"),

    # YEAR 8 - AUTUMN 2: BBC Micro:bit Programming
    ("Y8", "Autumn 2", 1, "Introduction to Micro:bit", "Identify Micro:bit components; Write first program", "Micro:bit, LED matrix, Buttons, Pins, MakeCode", "Micro:bit devices, MakeCode", "Explore MakeCode"),
    ("Y8", "Autumn 2", 2, "LED Display Programming", "Create images and animations; Use show and scroll", "LED, Pixel, Animation, Scroll, Show", "Display challenges", "Create custom image"),
    ("Y8", "Autumn 2", 3, "Buttons and Input", "Use button events; Create interactive programs", "Event, Button A/B, on button pressed", "Button challenges", "Button program"),
    ("Y8", "Autumn 2", 4, "Sensors: Accelerometer", "Use accelerometer data; Create motion-based programs", "Accelerometer, X/Y/Z axis, Gesture, Shake", "Motion challenges", "Motion-based program"),
    ("Y8", "Autumn 2", 5, "Variables and Selection", "Use variables with Micro:bit; Implement game logic", "Variable, Score, Lives, Selection, Game state", "Game development", "Micro:bit game"),
    ("Y8", "Autumn 2", 6, "Micro:bit Project", "Plan and create complete Micro:bit project", "Project planning, Development, Testing", "Project brief", "Complete project"),

    # YEAR 8 - SPRING 1: Algorithms: Searching & Sorting
    ("Y8", "Spring 1", 1, "Introduction to Algorithms", "Define algorithm; Understand importance of efficiency", "Algorithm, Efficiency, Steps, Order, Correctness", "Algorithm examples", "Write everyday algorithm"),
    ("Y8", "Spring 1", 2, "Linear Search", "Understand and trace linear search; Identify pros/cons", "Linear search, Sequential, Index, Found, Not found", "Search simulation", "Trace linear search"),
    ("Y8", "Spring 1", 3, "Binary Search", "Understand and trace binary search; Compare to linear", "Binary search, Midpoint, Divide, Sorted list", "Binary search cards", "Trace binary search"),
    ("Y8", "Spring 1", 4, "Bubble Sort", "Understand and trace bubble sort; Identify passes", "Bubble sort, Pass, Swap, Adjacent, Comparison", "Sorting cards activity", "Trace bubble sort"),
    ("Y8", "Spring 1", 5, "Merge Sort Concepts", "Understand merge sort principle; Compare sorting algorithms", "Merge sort, Divide and conquer, Merge, Efficiency", "Merge demonstration", "Compare algorithms"),
    ("Y8", "Spring 1", 6, "Algorithm Assessment", "Apply search and sort knowledge; Complete assessment", "All algorithm terms", "Assessment paper", "Revise algorithms"),

    # YEAR 8 - SPRING 2: Memory & Storage: Primary
    ("Y8", "Spring 2", 1, "Introduction to Computer Memory", "Understand memory role; Identify memory types", "Memory, Primary storage, Secondary storage, Capacity", "Memory diagram", "Memory types research"),
    ("Y8", "Spring 2", 2, "RAM: Random Access Memory", "Understand RAM purpose; Identify RAM characteristics", "RAM, Volatile, Read/write, Programs, Data", "RAM exploration", "RAM questions"),
    ("Y8", "Spring 2", 3, "ROM: Read Only Memory", "Understand ROM purpose; Compare RAM and ROM", "ROM, Non-volatile, Read only, BIOS, Firmware", "RAM vs ROM comparison", "Comparison table"),
    ("Y8", "Spring 2", 4, "Cache Memory", "Understand cache purpose; Explain memory hierarchy", "Cache, L1/L2/L3, Speed, CPU, Hierarchy", "Cache simulation", "Memory hierarchy diagram"),
    ("Y8", "Spring 2", 5, "Virtual Memory", "Understand virtual memory; Identify advantages/disadvantages", "Virtual memory, Paging, Swap, Disk thrashing", "Virtual memory demonstration", "Virtual memory questions"),
    ("Y8", "Spring 2", 6, "Memory Assessment", "Apply memory knowledge; Complete assessment", "All memory terms", "Assessment paper", "Revise memory"),

    # YEAR 8 - SUMMER 1: Secondary Storage
    ("Y8", "Summer 1", 1, "Need for Secondary Storage", "Explain why secondary storage is needed; Identify characteristics", "Secondary storage, Non-volatile, Permanent, Capacity", "Storage examples", "Storage at home audit"),
    ("Y8", "Summer 1", 2, "Magnetic Storage", "Understand how HDDs work; Identify pros/cons", "HDD, Platter, Read/write head, Magnetic, Sectors", "HDD demonstration", "HDD fact file"),
    ("Y8", "Summer 1", 3, "Optical Storage", "Understand how CDs/DVDs/Blu-ray work; Identify pros/cons", "Optical, CD, DVD, Blu-ray, Laser, Pits, Lands", "Optical media samples", "Optical storage comparison"),
    ("Y8", "Summer 1", 4, "Solid State Storage", "Understand how SSDs work; Compare to HDD", "SSD, Flash memory, NAND, No moving parts, USB", "SSD vs HDD comparison", "SSD advantages"),
    ("Y8", "Summer 1", 5, "Cloud Storage", "Understand cloud storage; Evaluate pros/cons", "Cloud, Remote server, Internet, Sync, Backup", "Cloud services comparison", "Cloud storage evaluation"),
    ("Y8", "Summer 1", 6, "Storage Assessment", "Compare all storage types; Complete assessment", "All storage terms, Capacity, Speed, Cost, Durability", "Assessment paper", "Revise storage"),

    # YEAR 8 - SUMMER 2: LEGO Robotics
    ("Y8", "Summer 2", 1, "Introduction to Robotics", "Understand robotics concepts; Identify robot components", "Robot, Sensor, Motor, Controller, Program", "Robot demonstration", "Research robots"),
    ("Y8", "Summer 2", 2, "Motors and Movement", "Program basic movement; Use motor blocks", "Motor, Power, Duration, Rotation, Direction", "Movement challenges", "Movement program"),
    ("Y8", "Summer 2", 3, "Sensors: Touch and Distance", "Use sensors for input; Create reactive programs", "Touch sensor, Ultrasonic, Distance, Trigger", "Sensor challenges", "Sensor program"),
    ("Y8", "Summer 2", 4, "Loops and Decision Making", "Use loops in robot programs; Implement decisions", "Loop, Wait, Switch, Condition", "Complex challenges", "Loop/decision program"),
    ("Y8", "Summer 2", 5, "Robot Challenge: Planning", "Plan solution to challenge; Design algorithm", "Decomposition, Planning, Testing, Iteration", "Challenge brief", "Plan solution"),
    ("Y8", "Summer 2", 6, "Robot Challenge: Competition", "Complete and test solution; Present to class", "Refinement, Optimization, Presentation", "Challenge materials", "Reflect on learning"),

    # YEAR 9 - AUTUMN 1: Advanced Micro:bit Projects
    ("Y9", "Autumn 1", 1, "Micro:bit Radio Communication", "Use radio to send messages between Micro:bits", "Radio, Send, Receive, Channel, Group", "Multiple Micro:bits", "Radio messaging program"),
    ("Y9", "Autumn 1", 2, "Data Transmission Project", "Create a data transmission system; Understand protocols", "Protocol, Data packet, Acknowledgment", "Project materials", "Develop transmission system"),
    ("Y9", "Autumn 1", 3, "Sensor Data Logging", "Log sensor data over time; Export and analyse", "Data logging, CSV, Timestamp, Analysis", "Data logging tools", "Log temperature data"),
    ("Y9", "Autumn 1", 4, "Python on Micro:bit", "Transition from MakeCode to Python; Write Python programs", "MicroPython, Syntax, Import, Module", "Mu editor, Python guide", "Python Micro:bit program"),
    ("Y9", "Autumn 1", 5, "Extended Project: Planning", "Plan a complex Micro:bit project; Define requirements", "Requirements, Scope, Timeline, Resources", "Project planning template", "Complete project plan"),
    ("Y9", "Autumn 1", 6, "Extended Project: Development", "Develop and test project; Document progress", "Development, Testing, Documentation, Iteration", "Project support", "Continue development"),

    # YEAR 9 - AUTUMN 2: Game Development with Python
    ("Y9", "Autumn 2", 1, "Introduction to Pygame", "Set up Pygame; Create game window; Understand game loop", "Pygame, Window, Game loop, FPS, Event", "Pygame documentation", "Basic window program"),
    ("Y9", "Autumn 2", 2, "Drawing and Colours", "Draw shapes; Use RGB colours; Create backgrounds", "Draw, Rect, Circle, RGB, Surface", "Drawing challenges", "Create game background"),
    ("Y9", "Autumn 2", 3, "User Input and Movement", "Handle keyboard/mouse input; Move objects", "Event handling, Key press, Mouse, Movement", "Input challenges", "Player movement program"),
    ("Y9", "Autumn 2", 4, "Collision Detection", "Detect collisions between objects; React to collisions", "Collision, Rect, Colliderect, Hitbox", "Collision challenges", "Add collision detection"),
    ("Y9", "Autumn 2", 5, "Sprites and Images", "Load and display images; Create sprite animations", "Sprite, Image, Load, Blit, Animation", "Sprite resources", "Add sprites to game"),
    ("Y9", "Autumn 2", 6, "Game Project: Development", "Develop complete game; Add scoring and sound", "Score, Sound, Game over, Win condition", "Project support", "Complete game project"),

    # YEAR 9 - SPRING 1: CPU Architecture
    ("Y9", "Spring 1", 1, "Introduction to CPU", "Identify CPU role; Understand basic components", "CPU, Processor, ALU, CU, Registers", "CPU diagram", "Label CPU diagram"),
    ("Y9", "Spring 1", 2, "The ALU and Control Unit", "Explain ALU and CU functions; Understand registers", "ALU, CU, Accumulator, MAR, MDR, PC", "Component cards", "Component functions"),
    ("Y9", "Spring 1", 3, "Fetch-Decode-Execute Cycle", "Trace the FDE cycle; Understand instruction processing", "Fetch, Decode, Execute, Cycle, Instruction", "FDE simulation", "Trace FDE cycle"),
    ("Y9", "Spring 1", 4, "CPU Performance Factors", "Explain clock speed, cores, cache impact", "Clock speed, GHz, Cores, Cache, Pipeline", "Comparison activity", "CPU comparison task"),
    ("Y9", "Spring 1", 5, "Von Neumann Architecture", "Describe Von Neumann architecture; Identify bottleneck", "Von Neumann, Stored program, Bus, Bottleneck", "Architecture diagram", "Architecture questions"),
    ("Y9", "Spring 1", 6, "CPU Assessment", "Apply CPU knowledge; Complete assessment", "All CPU terms", "Assessment paper", "Revise CPU"),

    # YEAR 9 - SPRING 2: Networks: Types & Topologies
    ("Y9", "Spring 2", 1, "Network Types: LAN and WAN", "Define and compare LAN and WAN; Give examples", "LAN, WAN, Local, Wide, Geographic", "Network examples", "Network types worksheet"),
    ("Y9", "Spring 2", 2, "Network Topologies: Star", "Describe star topology; Evaluate pros/cons", "Star topology, Central switch, Hub, Node", "Topology diagram", "Star topology analysis"),
    ("Y9", "Spring 2", 3, "Network Topologies: Mesh and Bus", "Describe mesh and bus; Compare topologies", "Mesh, Bus, Backbone, Full/partial mesh", "Topology comparison", "Topology comparison table"),
    ("Y9", "Spring 2", 4, "Wired vs Wireless Networks", "Compare wired and wireless; Understand technologies", "Ethernet, WiFi, Bandwidth, Interference, Security", "Comparison activity", "Wired vs wireless essay"),
    ("Y9", "Spring 2", 5, "Network Hardware", "Identify network hardware; Explain functions", "NIC, Switch, Router, WAP, Modem", "Hardware identification", "Hardware fact file"),
    ("Y9", "Spring 2", 6, "Networks Assessment", "Apply network knowledge; Complete assessment", "All network terms", "Assessment paper", "Revise networks"),

    # YEAR 9 - SUMMER 1: Networks: Protocols & Internet
    ("Y9", "Summer 1", 1, "Introduction to Protocols", "Define protocol; Understand need for standards", "Protocol, Standard, Communication, Rules", "Protocol examples", "Protocol research"),
    ("Y9", "Summer 1", 2, "TCP/IP Protocol Stack", "Understand TCP/IP layers; Identify layer functions", "TCP, IP, Layer, Packet, Encapsulation", "Layer diagram", "Layer functions"),
    ("Y9", "Summer 1", 3, "HTTP, HTTPS and FTP", "Understand web protocols; Explain secure connections", "HTTP, HTTPS, FTP, SSL/TLS, Encryption", "Protocol demonstration", "Protocol comparison"),
    ("Y9", "Summer 1", 4, "Email Protocols", "Understand SMTP, POP, IMAP; Explain email transmission", "SMTP, POP3, IMAP, Email server", "Email simulation", "Email protocols worksheet"),
    ("Y9", "Summer 1", 5, "IP Addresses and DNS", "Understand IP addressing; Explain DNS function", "IP address, IPv4, IPv6, DNS, Domain name", "DNS lookup activity", "IP and DNS questions"),
    ("Y9", "Summer 1", 6, "Packet Switching", "Understand packet switching; Trace packet journey", "Packet, Header, Payload, Route, Reassembly", "Packet simulation", "Packet switching trace"),

    # YEAR 9 - SUMMER 2: Python: Data Structures & File Handling
    ("Y9", "Summer 2", 1, "Lists in Python", "Create and manipulate lists; Access elements by index", "List, Index, Append, Remove, Length", "List challenges", "List practice programs"),
    ("Y9", "Summer 2", 2, "List Operations", "Iterate through lists; Use list methods", "Iteration, For loop, Sort, Reverse, Search", "List operations challenges", "List operations programs"),
    ("Y9", "Summer 2", 3, "2D Arrays/Lists", "Create and access 2D structures; Use nested loops", "2D array, Row, Column, Nested loop, Grid", "2D array challenges", "2D array programs"),
    ("Y9", "Summer 2", 4, "File Reading", "Open and read text files; Process file data", "File, Open, Read, Close, With statement", "File reading challenges", "File reading program"),
    ("Y9", "Summer 2", 5, "File Writing", "Write data to files; Append to files", "Write, Append, Mode, Newline", "File writing challenges", "File writing program"),
    ("Y9", "Summer 2", 6, "CSV Files & Assessment", "Read and write CSV files; Assessment", "CSV, Comma-separated, Split, Join", "CSV challenges, assessment", "Revise data structures"),

    # YEAR 10 (3 lessons per week - showing weekly topics with sub-lessons)
    # YEAR 10 - AUTUMN 1: Data Representation: Text & Images
    ("Y10", "Autumn 1", 1, "Character Sets: ASCII", "Understand ASCII encoding; Convert characters to/from codes", "ASCII, Character set, Encoding, 7-bit, 128 characters", "ASCII table, conversion practice", "ASCII conversions"),
    ("Y10", "Autumn 1", 2, "Character Sets: Unicode", "Understand Unicode; Compare ASCII and Unicode", "Unicode, UTF-8, UTF-16, Extended characters, Emoji", "Unicode exploration", "Character encoding questions"),
    ("Y10", "Autumn 1", 3, "Bitmap Image Basics", "Understand bitmap representation; Identify key terms", "Bitmap, Pixel, Resolution, Raster, Grid", "Pixel art activity", "Bitmap questions"),
    ("Y10", "Autumn 1", 4, "Colour Depth", "Calculate colour depth impact; Understand bit depth", "Colour depth, Bit depth, 1-bit, 8-bit, 24-bit", "Colour depth examples", "Colour depth calculations"),
    ("Y10", "Autumn 1", 5, "Image File Size Calculations", "Calculate image file sizes; Understand metadata", "File size, Resolution × Colour depth, Metadata", "Calculation practice", "File size problems"),
    ("Y10", "Autumn 1", 6, "Data Representation Review & Test", "Review text and image representation; Assessment", "All terms", "Assessment paper", "Revise data representation"),

    # YEAR 10 - AUTUMN 2: Data Representation: Sound & Compression
    ("Y10", "Autumn 2", 1, "Sound Sampling", "Understand digital sound; Explain sampling process", "Analogue, Digital, Sample, ADC", "Sound wave diagrams", "Sampling questions"),
    ("Y10", "Autumn 2", 2, "Sample Rate and Bit Depth", "Explain sample rate and bit depth; Calculate quality", "Sample rate, Hz, Bit depth, Quality", "Audio comparison", "Sample calculations"),
    ("Y10", "Autumn 2", 3, "Sound File Size Calculations", "Calculate sound file sizes", "File size = sample rate × bit depth × duration × channels", "Calculation practice", "Sound file size problems"),
    ("Y10", "Autumn 2", 4, "Compression: Lossy vs Lossless", "Compare compression types; Give examples", "Compression, Lossy, Lossless, MP3, PNG, ZIP", "Compression comparison", "Compression evaluation"),
    ("Y10", "Autumn 2", 5, "Run Length Encoding", "Apply RLE compression; Calculate compression ratio", "RLE, Run, Length, Ratio, Efficiency", "RLE practice", "RLE problems"),
    ("Y10", "Autumn 2", 6, "Compression Review & Test", "Review compression concepts; Assessment", "All compression terms", "Assessment paper", "Revise compression"),

    # YEAR 10 - SPRING 1: Programming: Subprograms & Testing
    ("Y10", "Spring 1", 1, "Functions in Python", "Define and call functions; Understand benefits", "Function, def, Call, Modular, Reusable", "Function challenges", "Create functions"),
    ("Y10", "Spring 1", 2, "Parameters and Arguments", "Use parameters; Pass arguments to functions", "Parameter, Argument, Passing values", "Parameter practice", "Functions with parameters"),
    ("Y10", "Spring 1", 3, "Return Values", "Return values from functions; Use returned data", "Return, Return value, Output", "Return challenges", "Return value programs"),
    ("Y10", "Spring 1", 4, "Variable Scope", "Understand local and global scope; Avoid scope errors", "Scope, Local, Global, Lifetime", "Scope demonstration", "Scope questions"),
    ("Y10", "Spring 1", 5, "Validation Techniques", "Implement input validation; Use appropriate checks", "Validation, Range check, Type check, Presence check, Length check", "Validation challenges", "Add validation to programs"),
    ("Y10", "Spring 1", 6, "Testing Strategies", "Create test plans; Test with normal, boundary, erroneous data", "Testing, Test plan, Normal, Boundary, Erroneous", "Test plan creation", "Create test plan"),

    # YEAR 10 - SPRING 2: Network Security & Threats
    ("Y10", "Spring 2", 1, "Malware Types", "Identify malware types; Explain how they spread", "Malware, Virus, Worm, Trojan, Ransomware, Spyware", "Malware case studies", "Malware fact file"),
    ("Y10", "Spring 2", 2, "Social Engineering", "Understand social engineering; Identify techniques", "Social engineering, Phishing, Pretexting, Baiting, Shoulder surfing", "Phishing examples", "Social engineering scenarios"),
    ("Y10", "Spring 2", 3, "Network Attacks", "Explain DDoS, brute force, SQL injection", "DDoS, Brute force, SQL injection, Data interception", "Attack demonstrations", "Attack explanations"),
    ("Y10", "Spring 2", 4, "Preventing Attacks: Firewalls", "Understand firewall function; Configure rules", "Firewall, Packet filtering, Rules, Ports", "Firewall configuration", "Firewall questions"),
    ("Y10", "Spring 2", 5, "Encryption", "Understand encryption; Compare symmetric/asymmetric", "Encryption, Decryption, Key, Symmetric, Asymmetric, Caesar cipher", "Encryption practice", "Encryption activity"),
    ("Y10", "Spring 2", 6, "Security Review & Test", "Review security measures; Assessment", "All security terms", "Assessment paper", "Revise security"),

    # YEAR 10 - SUMMER 1: Ethical, Legal & Environmental Issues
    ("Y10", "Summer 1", 1, "Data Protection Act", "Understand DPA principles; Apply to scenarios", "DPA, GDPR, Personal data, Principles, Rights", "DPA case studies", "DPA scenarios"),
    ("Y10", "Summer 1", 2, "Computer Misuse Act", "Understand CMA offences; Identify illegal activities", "CMA, Unauthorised access, Hacking, Offences", "CMA case studies", "CMA scenarios"),
    ("Y10", "Summer 1", 3, "Copyright and Licensing", "Understand copyright; Compare licence types", "Copyright, License, Open source, Proprietary, Creative Commons", "License comparison", "Licensing questions"),
    ("Y10", "Summer 1", 4, "Environmental Impact", "Evaluate environmental issues; Consider sustainability", "E-waste, Energy consumption, Carbon footprint, Sustainability", "Environmental audit", "Environmental impact essay"),
    ("Y10", "Summer 1", 5, "Ethical and Cultural Issues", "Discuss ethical implications; Consider cultural impact", "Ethics, Digital divide, Privacy, Automation, AI", "Ethical debates", "Ethical case study"),
    ("Y10", "Summer 1", 6, "Issues Review & Test", "Review legal and ethical issues; Assessment", "All terms", "Assessment paper", "Revise issues"),

    # YEAR 10 - SUMMER 2: NEA Preparation & Programming Practice
    ("Y10", "Summer 2", 1, "NEA Overview", "Understand NEA requirements; Review assessment criteria", "NEA, Programming task, Assessment criteria, Marks", "NEA guidance", "Review NEA requirements"),
    ("Y10", "Summer 2", 2, "Programming Task Practice 1", "Complete practice programming task under exam conditions", "Time management, Problem solving, Decomposition", "Practice task", "Complete practice task"),
    ("Y10", "Summer 2", 3, "Programming Task Review", "Review practice task; Identify improvements", "Self-assessment, Peer review, Feedback", "Review materials", "Reflect on practice"),
    ("Y10", "Summer 2", 4, "Advanced Programming Techniques", "Apply string methods, list comprehension, error handling", "String methods, List comprehension, Try-except", "Advanced challenges", "Advanced programs"),
    ("Y10", "Summer 2", 5, "Exam-Style Questions", "Practice exam-style programming questions", "Exam technique, Time management, Mark schemes", "Past papers", "Exam questions"),
    ("Y10", "Summer 2", 6, "End of Year Assessment", "Complete end of year assessment", "All Y10 content", "Assessment paper", "Summer revision"),

    # YEAR 11 - AUTUMN 1: NEA Completion
    ("Y11", "Autumn 1", 1, "NEA: Understanding the Task", "Analyse NEA task; Plan approach", "Task analysis, Requirements, Planning", "NEA task paper", "Begin planning"),
    ("Y11", "Autumn 1", 2, "NEA: Initial Development", "Begin coding solution; Implement core features", "Development, Core functionality, Iteration", "Development support", "Continue development"),
    ("Y11", "Autumn 1", 3, "NEA: Feature Development", "Implement additional features; Add validation", "Features, Validation, Robustness", "Development support", "Continue development"),
    ("Y11", "Autumn 1", 4, "NEA: Testing", "Test program thoroughly; Document testing", "Testing, Test plan, Evidence, Screenshots", "Testing template", "Complete testing"),
    ("Y11", "Autumn 1", 5, "NEA: Refinement", "Refine code; Add comments and documentation", "Refinement, Comments, Documentation", "Documentation guide", "Complete refinement"),
    ("Y11", "Autumn 1", 6, "NEA: Final Submission", "Complete final checks; Submit NEA", "Submission, Final check, Deadline", "Submission checklist", "Submit NEA"),

    # YEAR 11 - AUTUMN 2: Computational Thinking Review
    ("Y11", "Autumn 2", 1, "Decomposition and Abstraction", "Review decomposition; Apply abstraction", "Decomposition, Abstraction, Sub-problems, Simplification", "Problem-solving activities", "Decomposition practice"),
    ("Y11", "Autumn 2", 2, "Algorithm Design", "Design algorithms for problems; Use pseudocode", "Algorithm, Pseudocode, Efficiency, Steps", "Algorithm challenges", "Design algorithms"),
    ("Y11", "Autumn 2", 3, "Trace Tables", "Complete trace tables accurately; Track variable values", "Trace table, Variable, Iteration, Step-through", "Trace table practice", "Trace table exercises"),
    ("Y11", "Autumn 2", 4, "Search Algorithms Review", "Review linear and binary search; Compare efficiency", "Linear search, Binary search, Efficiency, Big O", "Search algorithm practice", "Search questions"),
    ("Y11", "Autumn 2", 5, "Sort Algorithms Review", "Review bubble and merge sort; Trace sorting", "Bubble sort, Merge sort, Pass, Comparison", "Sort algorithm practice", "Sort questions"),
    ("Y11", "Autumn 2", 6, "Computational Thinking Test", "Complete topic assessment", "All computational thinking terms", "Assessment paper", "Revise topic"),

    # YEAR 11 - SPRING 1: Systems Architecture Review
    ("Y11", "Spring 1", 1, "CPU Architecture Review", "Review CPU components; Explain FDE cycle", "CPU, ALU, CU, Registers, FDE", "CPU diagrams", "CPU questions"),
    ("Y11", "Spring 1", 2, "Memory Review", "Review RAM, ROM, cache, virtual memory", "RAM, ROM, Cache, Virtual memory", "Memory comparison", "Memory questions"),
    ("Y11", "Spring 1", 3, "Storage Review", "Review storage types; Compare technologies", "HDD, SSD, Optical, Cloud, Magnetic", "Storage comparison", "Storage questions"),
    ("Y11", "Spring 1", 4, "Embedded Systems", "Review embedded systems; Give examples", "Embedded system, Dedicated function, Real-time", "Embedded examples", "Embedded questions"),
    ("Y11", "Spring 1", 5, "Systems Past Paper Practice", "Complete past paper questions on systems", "Exam technique, Mark allocation", "Past papers", "Past paper questions"),
    ("Y11", "Spring 1", 6, "Systems Architecture Test", "Complete topic assessment", "All systems terms", "Assessment paper", "Revise topic"),

    # YEAR 11 - SPRING 2: Networks Review
    ("Y11", "Spring 2", 1, "Network Types and Topologies Review", "Review LAN, WAN, topologies", "LAN, WAN, Star, Mesh, Bus", "Network diagrams", "Network questions"),
    ("Y11", "Spring 2", 2, "Protocols Review", "Review TCP/IP, HTTP, FTP, email protocols", "TCP/IP, HTTP, HTTPS, FTP, SMTP, POP, IMAP", "Protocol practice", "Protocol questions"),
    ("Y11", "Spring 2", 3, "Internet and Addressing Review", "Review IP, DNS, packet switching", "IP address, DNS, Packet switching", "Internet simulation", "Internet questions"),
    ("Y11", "Spring 2", 4, "Network Security Review", "Review threats and security measures", "Malware, Firewall, Encryption, Social engineering", "Security scenarios", "Security questions"),
    ("Y11", "Spring 2", 5, "Networks Past Paper Practice", "Complete past paper questions on networks", "Exam technique, Mark allocation", "Past papers", "Past paper questions"),
    ("Y11", "Spring 2", 6, "Networks Test", "Complete topic assessment", "All network terms", "Assessment paper", "Revise topic"),

    # YEAR 11 - SUMMER 1: Final Revision
    ("Y11", "Summer 1", 1, "Data Representation Revision", "Comprehensive binary, hex, text, image, sound revision", "Binary, Hex, ASCII, Unicode, Images, Sound, Compression", "Revision materials", "Data questions"),
    ("Y11", "Summer 1", 2, "Programming Revision", "Review all programming concepts; Practice coding", "Variables, Selection, Iteration, Functions, Data structures, Files", "Programming challenges", "Programming practice"),
    ("Y11", "Summer 1", 3, "Legal and Ethical Revision", "Review laws and ethical issues", "DPA, CMA, Copyright, Ethics, Environment", "Case studies", "Issues questions"),
    ("Y11", "Summer 1", 4, "Mock Exam: Paper 1", "Complete full Paper 1 under exam conditions", "Exam technique, Time management", "Mock paper", "Review mock"),
    ("Y11", "Summer 1", 5, "Mock Exam: Paper 2", "Complete full Paper 2 under exam conditions", "Programming, Problem solving", "Mock paper", "Review mock"),
    ("Y11", "Summer 1", 6, "Mock Review and Final Prep", "Review mocks; Target weak areas", "Self-assessment, Target setting", "Mock feedback", "Final revision"),

    # YEAR 11 - SUMMER 2: Examinations
    ("Y11", "Summer 2", 1, "Final Revision: Paper 1 Focus", "Intensive Paper 1 revision; Past papers", "All Paper 1 topics", "Past papers", "Past papers"),
    ("Y11", "Summer 2", 2, "Final Revision: Paper 2 Focus", "Intensive Paper 2 revision; Programming practice", "All Paper 2 topics", "Past papers, practice tasks", "Practice tasks"),
    ("Y11", "Summer 2", 3, "Exam Preparation", "Final exam tips; Stress management; Equipment check", "Exam technique, Wellbeing", "Exam guidance", "Rest and prepare"),
    ("Y11", "Summer 2", 4, "GCSE Paper 1", "Principles of Computer Science examination", "Paper 1", "Exam", ""),
    ("Y11", "Summer 2", 5, "GCSE Paper 2", "Application of Computational Thinking examination", "Paper 2", "Exam", ""),
    ("Y11", "Summer 2", 6, "Post-Exam", "Course completion; Feedback; Next steps", "Reflection", "Course evaluation", ""),
]

# Populate medium term data
for row_idx, data in enumerate(medium_term_data, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_medium.cell(row=row_idx, column=col_idx, value=value)
        apply_cell_style(cell, data[0])

# Set column widths
medium_widths = [6, 12, 6, 40, 70, 55, 40, 35]
for col, width in enumerate(medium_widths, 1):
    ws_medium.column_dimensions[get_column_letter(col)].width = width

ws_medium.freeze_panes = 'A2'

# ============================================================================
# SHEET 3: SHORT TERM PLAN (By Lessons) - Sample for first half-term of each year
# ============================================================================
ws_short = wb.create_sheet("Short Term Plan")

# Headers
short_headers = ["Year", "Half-Term", "Week", "Lesson", "Learning Objective", "Starter (10 min)", "Main Activity (35-40 min)", "Plenary (10 min)", "Differentiation", "Resources", "Assessment"]
for col, header in enumerate(short_headers, 1):
    ws_short.cell(row=1, column=col, value=header)
apply_header_style(ws_short)

# Short term data (lesson by lesson) - Sample lessons
short_term_data = [
    # Y7 Autumn 1 Week 1 - What is the Internet? (1 lesson per week KS3)
    ("Y7", "Autumn 1", 1, 1, "Understand the difference between the internet and the World Wide Web",
     "Discussion: What did you use the internet for this week? List devices that connect.",
     "Input: Teacher presentation on internet vs web, history of ARPANET. Activity: Students create a network diagram showing how their home connects to the internet. Students label: device, router, ISP, servers.",
     "Quick quiz: 5 questions on internet vs web definitions. Exit ticket: One thing learned.",
     "Support: Diagram template provided. Extend: Research Tim Berners-Lee",
     "Presentation slides, network diagram worksheet, mini whiteboards",
     "Formative: Diagram accuracy, exit ticket"),

    # Y7 Autumn 1 Week 2 - Online Safety
    ("Y7", "Autumn 1", 2, 1, "Identify online risks and understand how to protect personal information",
     "Scenario cards: Students sort scenarios into 'safe' and 'risky' online behaviour.",
     "Input: CEOP video on online safety. Discussion: What personal information should never be shared? Activity: Students create a 'Personal Information Audit' - list what they share online and evaluate risks. Role-play: Responding to suspicious messages.",
     "Create class 'Top 5 Online Safety Rules' poster. Peer discussion of key takeaways.",
     "Support: Scaffolded audit sheet. Extend: Research cases of identity theft",
     "CEOP video, scenario cards, audit worksheet, poster paper",
     "Formative: Participation in discussions, audit completion"),

    # Y7 Autumn 2 Week 1 - Introduction to Binary
    ("Y7", "Autumn 2", 1, 1, "Understand why computers use binary and convert small binary numbers to denary",
     "Puzzle: Use only ON/OFF switches to represent numbers. How would you show 0-3?",
     "Input: Explain binary as computer's language, transistors, bits. Demo: Binary cards activity (1, 2, 4, 8, 16 cards). Students work in pairs converting 4-bit binary to denary. Progress to individual practice with worksheet.",
     "Binary finger counting game - show numbers 0-31 on one hand. Review misconceptions.",
     "Support: Place value mat with worked examples. Extend: 8-bit conversions",
     "Binary cards (sets for each pair), place value charts, conversion worksheet",
     "Formative: Worksheet completion, binary finger accuracy"),

    # Y7 Spring 1 Week 1 - Introduction to Edublocks
    ("Y7", "Spring 1", 1, 1, "Navigate the Edublocks interface and execute basic turtle commands",
     "Discussion: What is programming? Share ideas on what makes a good set of instructions.",
     "Demo: Teacher shows Edublocks interface, explains blocks and how to run code. Task 1: Students log in and explore interface (5 min). Task 2: Follow instructions to draw a straight line. Task 3: Draw a square using forward and turn blocks. Extension: Experiment with colours and pen sizes.",
     "Gallery walk: Students view each other's squares. Discuss: What commands did everyone use?",
     "Support: Step-by-step instruction card. Extend: Create a rectangle, triangle",
     "Computers with Edublocks access, instruction cards, success criteria checklist",
     "Formative: Square completion, peer feedback"),

    # Y8 Autumn 1 Week 1 - Introduction to Logic Gates
    ("Y8", "Autumn 1", 1, 1, "Understand Boolean logic in hardware and identify AND gate behaviour",
     "Puzzle: Light switches - if BOTH switches must be ON for light, draw the circuit.",
     "Input: Introduce logic gates as decision-makers in circuits. Explain AND gate with truth table. Demo: Online logic gate simulator. Activity: Students complete AND gate truth table, then design scenarios (e.g., alarm system: door AND window must be open).",
     "Mini whiteboard quiz: Show AND gate truth table outputs for given inputs.",
     "Support: Partially completed truth table. Extend: Research where AND gates are used",
     "Logic gate simulator (online), truth table worksheets, mini whiteboards",
     "Formative: Truth table accuracy, scenario design"),

    # Y8 Autumn 2 Week 1 - Introduction to Micro:bit
    ("Y8", "Autumn 2", 1, 1, "Identify Micro:bit components and write a first program to display text",
     "Explore: Pass around Micro:bits. What components can you identify? Make a list.",
     "Input: Label Micro:bit diagram (LEDs, buttons, pins, accelerometer, etc.). Demo: MakeCode interface and how to download code. Task 1: Display your name scrolling across LEDs. Task 2: Display a custom image/icon. Task 3: Make image change when button pressed.",
     "Share: Students show their displays. Vote on most creative. Discussion: What else could we program?",
     "Support: Guided MakeCode tutorial. Extend: Add sound or animation",
     "Micro:bits, USB cables, computers with MakeCode, component diagram worksheet",
     "Formative: Working name display, component identification"),

    # Y9 Autumn 1 Week 1 - Micro:bit Radio Communication
    ("Y9", "Autumn 1", 1, 1, "Use radio blocks to send messages between Micro:bits",
     "Discussion: How do devices communicate wirelessly? Examples from everyday life.",
     "Input: Explain Micro:bit radio feature, channels, groups. Demo: Simple send/receive program. Task: In pairs, set up radio channel. Person A sends message on button A press, Person B receives and displays. Swap roles. Extension: Create a simple chat system with different messages for each button.",
     "Challenge: Class-wide communication chain - pass a message around the room via radio.",
     "Support: Pre-written code to modify. Extend: Add encryption to messages",
     "Micro:bits (2 per pair), MakeCode, radio communication guide",
     "Formative: Successful message transmission, peer assessment"),

    # Y10 sample lessons (3 per week)
    ("Y10", "Autumn 1", 1, 1, "Understand ASCII encoding and convert characters to/from ASCII codes",
     "Decode this message: 72 69 76 76 79 (using ASCII table). What does it say?",
     "Input: Explain character encoding need, ASCII development, 7-bit = 128 characters. Students use ASCII table to encode their name. Worked examples of character-to-code and code-to-character. Practice worksheet with progressive difficulty.",
     "Peer marking of worksheet. Discussion: Limitations of ASCII?",
     "Support: Simplified ASCII table (A-Z, 0-9 only). Extend: Binary representation of ASCII",
     "ASCII tables, encoding worksheet, mini whiteboards",
     "Formative: Worksheet accuracy, peer marking"),

    ("Y10", "Autumn 1", 1, 2, "Understand extended character sets including Unicode",
     "Question: How would ASCII represent Chinese characters? Emoji? Why is this a problem?",
     "Input: Explain Unicode development, UTF-8/UTF-16, extended character support. Compare ASCII (128) vs Unicode (143,000+). Activity: Use Unicode table to find codes for various symbols and emojis. Calculate storage comparison: ASCII vs Unicode for same text.",
     "Exit ticket: Why was Unicode developed? Give 2 advantages over ASCII.",
     "Support: Guided comparison table. Extend: Research UTF-8 variable-width encoding",
     "Unicode reference, comparison activity sheet, calculators",
     "Formative: Exit ticket responses, comparison accuracy"),

    ("Y10", "Autumn 1", 1, 3, "Apply character encoding knowledge to exam-style questions",
     "Quick fire: Convert 'CAT' to ASCII codes (without table). Check with table.",
     "Exam technique: Review how character encoding questions appear. Mark scheme analysis. Students complete 3 exam-style questions independently. Peer mark using mark scheme. Address common misconceptions and errors.",
     "Traffic light self-assessment: Red/Amber/Green confidence on character encoding.",
     "Support: Worked example for each question type. Extend: Create own exam question with mark scheme",
     "Past paper questions, mark schemes, traffic light cards",
     "Formative: Question accuracy, self-assessment"),

    # Y11 sample lessons
    ("Y11", "Autumn 1", 1, 1, "Analyse the NEA task and begin planning the solution approach",
     "Review: NEA assessment criteria - what gains marks? Discuss with partner.",
     "Input: Distribute NEA task. Read through as class, highlight key requirements. Decomposition activity: Break task into sub-problems. Students begin individual planning: identify inputs, outputs, processes. Start pseudocode for main algorithm.",
     "Share plans with partner. Identify any missing requirements. Set targets for next lesson.",
     "Support: Planning template with prompts. Extend: Consider additional features beyond requirements",
     "NEA task sheet, planning template, assessment criteria, pseudocode guide",
     "Formative: Planning document, peer feedback"),

    ("Y11", "Autumn 1", 1, 2, "Begin development of core NEA functionality",
     "Review plan from last lesson. Prioritise: What must be coded first?",
     "Development time: Students work on core functionality. Teacher circulates providing 1:1 support. Checkpoint at 20 minutes: All students should have basic input/output working. Students document progress and any issues in development log.",
     "Progress share: 3 volunteers explain current progress and next steps. Homework: Continue development.",
     "Support: Starter code scaffold. Extend: Begin implementing validation",
     "Computers with Python, NEA task, development log template",
     "Formative: Code review, development log"),

    ("Y11", "Autumn 1", 1, 3, "Continue NEA development with focus on functionality",
     "Bug fixing warm-up: Fix deliberate errors in sample code (2 minutes).",
     "Continued development time. Focus: Core functionality must be working by end of lesson. Teacher 1:1 support rotation. Mid-lesson check: Students test their code with provided test data. Record results in testing log. Debug any failures.",
     "Exit ticket: Screenshot of working core functionality OR list of specific issues needing help.",
     "Support: Additional worked examples. Extend: Implement additional features",
     "Computers with Python, test data set, testing log template",
     "Formative: Working code evidence, testing log"),
]

# Populate short term data
for row_idx, data in enumerate(short_term_data, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_short.cell(row=row_idx, column=col_idx, value=value)
        apply_cell_style(cell, data[0])

# Set column widths
short_widths = [6, 12, 6, 7, 60, 50, 80, 50, 45, 50, 35]
for col, width in enumerate(short_widths, 1):
    ws_short.column_dimensions[get_column_letter(col)].width = width

ws_short.freeze_panes = 'A2'

# Set row heights for better readability
for ws in [ws_long, ws_medium, ws_short]:
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 45

# Save workbook
output_path = "/home/user/SoL/CS_Scheme_of_Learning_Y7-Y11_Edexcel_GCSE.xlsx"
wb.save(output_path)
print(f"Scheme of Learning saved to: {output_path}")
print("Worksheets created:")
print("1. Long Term Plan (by half-terms)")
print("2. Medium Term Plan (by weeks)")
print("3. Short Term Plan (by lessons - sample)")
